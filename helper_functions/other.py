import brian2.monitors
import numpy as np
import matplotlib.pyplot as plt
import os, datetime

"""
@author: t.f.tiotto@rug.nl
@university: University of Groningen
@group: Cognitive Modelling

Function:
- Has neuron spiked in the supplied time window?

Script arguments:
-

Script output:
- has_spiked: dictionary with neurons as keys and True or False depending if the neuron has spiked or not within the
time window

Comments:
- 

Inputs:
- window(tuple, list, Numpy array, or int, float): 2-element list-like object containing endpoints of time window,
if not list-like then the lower bound is assumed t=0
- data_monitor(brian2.data_monitor.SpikeMonitor): Brian2 SpikeMonitor object with recorded spikes
"""


def has_spiked(window, monitor):
    assert isinstance(monitor, brian2.monitors.SpikeMonitor)
    assert isinstance(window, brian2.Quantity)
    assert window.size == 1 or window.size == 2
    if window.size == 1:
        window = (0, window)

    spikes = monitor.spike_trains()
    has_spiked = np.zeros(len(spikes), dtype=bool)
    for i, spks in spikes.items():
        sp = spks[(spks > window[0]) & (spks < window[1])]
        # print(i, np.count_nonzero(sp))
        has_spiked[i] = bool(np.count_nonzero(sp))

    return has_spiked


def visualise_connectivity(S):
    Ns = len(S.source)
    Nt = len(S.target)
    plt.figure(figsize=(10, 4))
    plt.subplot(121)
    plt.plot(np.zeros(Ns), np.arange(Ns), 'ok', ms=10)
    plt.plot(np.ones(Nt), np.arange(Nt), 'ok', ms=10)
    for i, j in zip(S.i, S.j):
        plt.plot([0, 1], [i, j], '-k')
    plt.xticks([0, 1], ['Source', 'Target'])
    plt.ylabel('Neuron index')
    plt.xlim(-0.1, 1.1)
    plt.ylim(-1, max(Ns, Nt))
    plt.subplot(122)
    plt.plot(S.i, S.j, 'ok')
    plt.xlim(-1, Ns)
    plt.ylim(-1, Nt)
    plt.xlabel('Source neuron index')
    plt.ylabel('Target neuron index')
    plt.show()


def make_folders(path):
    if not os.path.exists(path):
        os.makedirs(path)


def make_timestamped_folder(path):
    if not os.path.exists(path):
        os.makedirs(path)
    timestamp = datetime.datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    os.makedirs(os.path.join(path, timestamp))

    return os.path.join(path, timestamp)


def contiguous_regions(condition):
    """Finds contiguous True regions of the boolean array "condition". Returns
    a 2D array where the first column is the start index of the region and the
    second column is the end index."""

    # Find the indices of changes in "condition"
    d = np.diff(condition)
    idx, = d.nonzero()

    # We need to start things after the change in "condition". Therefore,
    # we'll shift the index by 1 to the right.
    idx += 1

    if condition[0]:
        # If the start of condition is True prepend a 0
        idx = np.r_[0, idx]

    if condition[-1]:
        # If the end of condition is True, append the length of the array
        idx = np.r_[idx, condition.size - 1]  # Edit

    # Reshape the result into two columns
    idx.shape = (-1, 2)

    return idx


def append_df_to_excel(df, excel_path, sheet_name, index=False):
    import pandas as pd
    from openpyxl import load_workbook

    book = load_workbook(excel_path)
    if not sheet_name in book.sheetnames:
        book.create_sheet(sheet_name)
    book.save(excel_path)

    df_excel = pd.read_excel(excel_path, engine='openpyxl', sheet_name=sheet_name)
    result = pd.concat([df_excel, df], ignore_index=not index)

    excel_book = load_workbook(excel_path)
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        writer.book = excel_book
        writer.sheets = dict((ws.title, ws) for ws in excel_book.worksheets)

        result.to_excel(writer, index=index, sheet_name=sheet_name, engine='openpyxl')


def find_ps(path, sim_time, attractor, write_to_file=False, parameters=None, verbose=False):
    import pyspike as spk
    from scipy.ndimage.filters import uniform_filter1d
    import pandas as pd

    spike_trains = spk.load_spike_trains_from_txt(os.path.join(path, 'spikes_pyspike.txt'),
                                                  edges=(0, sim_time),
                                                  ignore_empty_lines=False)

    spike_sync_profile = spk.spike_sync_profile(spike_trains, indices=attractor[1])

    x, y = spike_sync_profile.get_plottable_data()
    # mean_filter_size = round( len( x ) / 10 )
    mean_filter_size = 20

    try:
        y_smooth = uniform_filter1d(y, size=mean_filter_size)
    except:
        y_smooth = np.zeros(len(x))

    # if there are no spikes we need to force it to count zero PSs
    if np.count_nonzero([i for i in spike_trains]) == 0:
        pss = np.array([[]])
    else:
        pss = contiguous_regions(y_smooth > 0.8)

    if write_to_file:
        assert parameters is not None

        parameter_names = list(parameters.keys())
        parameter_values = list(parameters.values())

        df = pd.DataFrame(columns=['atr'] +
                                  list(parameter_names) +
                                  ['start_s', 'end_s', 'center', 'max', 'mean', 'std'])

        if pss.size:
            for ps in pss:
                df = pd.concat([df, pd.DataFrame([[attractor[0]] +
                                                  parameter_values +
                                                  [x[ps[0]],
                                                   x[ps[1]],
                                                   x[ps[0] + np.argmax(y_smooth[ps[0]:ps[1]])],
                                                   np.max(y_smooth[ps[0]:ps[1]]),
                                                   np.mean(y_smooth[ps[0]:ps[1]]),
                                                   np.std(y_smooth[ps[0]:ps[1]])]],
                                                 columns=['atr'] +
                                                         list(parameter_names) +
                                                         ['start_s', 'end_s', 'center', 'max', 'mean', 'std'])])

        fn = os.path.join(path, "pss.xlsx")

        ensure_excel_exists(fn)

        append_df_to_excel(df, fn, sheet_name="PSs")

    if verbose:
        print(f'Found PS in {attractor[0]} '
              f'between {x[ps[0]]} s and {x[ps[1]]} s '
              f'centered at {x[ps[0] + np.argmax(y_smooth[ps[0]:ps[1]])]} s '
              f'with max value {np.max(y_smooth[ps[0]:ps[1]])} '
              f'(mean={np.mean(y_smooth[ps[0]:ps[1]])}, '
              f'std={np.std(y_smooth[ps[0]:ps[1]])}'
              )

    return x, y, y_smooth, pss


def count_pss_in_gss(pss_path, normalise_by_PS=False, num_gss=None, write_to_file=False, ba=None, gss=None,
                     verbose=True):
    import pandas as pd

    fn_pss = os.path.join(pss_path, 'pss.xlsx')
    df = pd.read_excel(fn_pss, engine='openpyxl', sheet_name='PSs')

    for atr in df['atr'].unique():
        num_ps_in_gs = 0
        for _, row in df.iterrows():
            if row['atr'] == atr:
                for gs in gss:
                    if row['start_s'] >= gs[1][0] and row['end_s'] <= gs[1][1]:
                        num_ps_in_gs += 1

        if normalise_by_PS:
            total_num = len(df)
            try:
                ps_in_gs = num_ps_in_gs / total_num * 100
            except ZeroDivisionError:
                ps_in_gs = 0
        else:
            assert num_gss is not None
            total_num = num_gss
            try:
                ps_in_gs = num_ps_in_gs / num_gss * 100
            except ZeroDivisionError:
                ps_in_gs = 0

        if verbose:
            print(f'In {atr}: found {num_ps_in_gs} PS in GS out of {total_num} total ({ps_in_gs} %)')

        if write_to_file:
            ensure_excel_exists(fn_pss)

            df_to_file = pd.DataFrame([[atr, ba, gss[0][0], num_ps_in_gs, total_num, ps_in_gs]],
                                      columns=['atr', 'ba_Hz', 'gs_%', 'num_ps_in_gs', 'total_num',
                                               'percent_ps_in_gs'])

            if normalise_by_PS:
                append_df_to_excel(df_to_file, fn_pss, sheet_name='PSs_in_GSs_norm_PS')
            else:
                append_df_to_excel(df_to_file, fn_pss, sheet_name='PSs_in_GSs_norm_GS')


def append_pss_to_xlsx(experiment_path, iteration_path):
    import pandas as pd

    fn_iteration = os.path.join(iteration_path, 'pss.xlsx')
    fn_experiment = os.path.join(experiment_path, 'pss.xlsx')

    ensure_excel_exists(fn_experiment)

    xls = pd.ExcelFile(fn_iteration)
    for sheet in xls.sheet_names:
        df = pd.read_excel(fn_iteration, sheet_name=sheet, engine='openpyxl')
        append_df_to_excel(df, fn_experiment, sheet_name=sheet)


def ensure_excel_exists(fn):
    import pandas as pd

    if not os.path.isfile(fn):
        with pd.ExcelWriter(fn, engine='openpyxl') as writer:
            df_empty = pd.DataFrame()
            df_empty.to_excel(writer, index=False, sheet_name="PSs")
            # df_empty.to_excel( writer, index=False, sheet_name="PSs_in_GSs" )


def compute_pss_statistics(timestamp_folder, parameters=None, write_to_file=True, verbose=True):
    import pandas as pd

    fn = os.path.join(timestamp_folder, 'pss.xlsx')
    df = pd.read_excel(fn, engine='openpyxl', sheet_name='PSs')

    group_pss_by_attractor = pd.DataFrame(df.groupby(by='atr').size(), columns=['Count'])
    group_pss_by_experiment = pd.DataFrame(df.groupby(parameters).size(), columns=['Count'])
    group_pss_by_experiment_and_attractor = pd.DataFrame(df.groupby(parameters + ['atr']).size(),
                                                         columns=['Count'])

    if verbose:
        print('PSs per attractor across whole experiment\n', group_pss_by_attractor)
        print('PSs within each experiment\n', group_pss_by_experiment)

    if write_to_file:
        ensure_excel_exists(fn)
        # TODO include zero counts in Excel file
        append_df_to_excel(group_pss_by_attractor, fn, sheet_name='PSs_groupby_atr', index=True)
        append_df_to_excel(group_pss_by_experiment, fn, sheet_name='PSs_groupby_exp', index=True)
        append_df_to_excel(group_pss_by_experiment_and_attractor, fn, sheet_name='PSs_groupby_exp_and_atr',
                           index=True)

    return df


def generate_gss(gs_percentage, gs_length, pre_runtime, gs_runtime, target=None, length=None):
    if length is None:
        length = gs_runtime - gs_length

    return [(gs_percentage, target, (pre_runtime, pre_runtime + gs_length), length)]


def generate_periodic_gss(gs_percentage, gs_freq, gs_length, pre_runtime, gs_runtime, target):
    import numpy as np
    from scipy import signal as sg

    cycle = 1 / gs_freq
    duty_perc = gs_length / cycle

    t = np.linspace(0, gs_runtime, 1000)

    signal = sg.square(2 * np.pi * gs_freq * t, duty=duty_perc)
    signal = np.where(signal > 0, 1, 0)

    switching_idx = np.where(np.abs(np.diff(signal)) > 0)[0]
    gss_times = t[switching_idx]
    gss_times = np.delete(gss_times, 0)
    gss_times = np.round(gss_times, decimals=2)

    gss = []
    for i in range(0, len(gss_times), 2):
        if isinstance(target, int):
            n_act_ids = int(target * gs_percentage / 100)
            act_ids = np.random.choice(target, n_act_ids, replace=False)
        try:
            gss.append(
                (gs_percentage, act_ids, (gss_times[i], gss_times[i + 1]),
                 gss_times[i + 2] - gss_times[i + 1])
            )
        except IndexError:
            pass

    return gss


def compile_xlsx_from_folders(base_folder):
    from os import listdir
    from os.path import isfile, join
    from natsort import os_sorted
    from tqdm import tqdm

    # base_fol = '/Users/thomas/PycharmProjects/wm_colaboration/results/RCN_attractor_reactivation/15-03-2022_21-15-09'
    iter_folds = os_sorted([f for f in listdir(base_folder) if not isfile(join(base_folder, f))])
    for fol in tqdm(iter_folds):
        append_pss_to_xlsx(base_folder, os.path.join(base_folder, fol))

    compute_pss_statistics(base_folder)


def remove_pickles(base_folder):
    import os

    count = 0
    for path, subdirs, files in os.walk(base_folder):
        for name in files:
            if name.endswith('.pickle'):
                os.remove(os.path.join(path, name))
                count += 1

    print(f'Removed {count} .pickle files')


def compile_overlapping_gss(gss1, gss2):
    """
    For now we assume that gss2 always "wins" when there's an overlap
    :param gss1:
    :param gss2:
    :return:
    """

    def overlap(a, b):
        return a[0] <= b[0] <= a[1] or b[0] <= a[0] <= b[1]

    for i, gs1 in enumerate(gss1):
        for j, gs2 in enumerate(gss2):
            if overlap(gs1[2], gs2[2]):
                gs2 = list(gs2)
                gs2[3] = gs1[3]
                gs2 = tuple(gs2)
                gss1[i] = gs2
            elif gs2 not in gss1:
                gs2 = list(gs2)
                gs2[3] = abs(gss1[i][2][0] - gs2[2][1])
                gs2 = tuple(gs2)
                gss1.insert(i, gs2)

    return gss1


def clear_screen():
    from os import system, name

    # for windows
    if name == 'nt':
        _ = system('cls')

    # for mac and linux(here, os.name is 'posix')
    else:
        _ = system('clear')
